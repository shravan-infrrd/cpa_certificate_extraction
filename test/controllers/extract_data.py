import logging

from flask import Flask, request
from flask import jsonify
from flask_restful import Resource

from controllers.scanned_to_machined import read_scanned_pdf, read_scanned_image
from exceptions.exceptions_handler import *
from utils import formulate_response, is_machine_generated
from constant import PDF_UPLOAD_DIRECTORY, PROJECT_ROOT, REFERENCE_FILE
from os import path
import os
from service.abby_data_extractor import extract_to_docx
import subprocess


from lib.parse_data import parse_all_fields
from openpyxl import Workbook
import openpyxl
import uuid

import spacy
from spacy import displacy
import en_core_web_sm


import json
from flask_pymongo import PyMongo
import copy


app = Flask('mongo')
app.config['MONGO_DBNAME'] = 'cpa_database'
app.config['MONGO_URI'] = 'mongodb://localhost:27017/cpa_database'

mongo = PyMongo(app)

class ExtractData(Resource):

    def update_excel_sheet(self, result, name):
        if os.path.exists(REFERENCE_FILE):
            wb = openpyxl.load_workbook( REFERENCE_FILE )
        else:
            wb = Workbook()
        try:
            sheet = wb[name]
        except:
            sheet = wb.create_sheet(name)
       
        for index, (key, value) in enumerate(result.items()):
            if index < 4:
                continue
            sheet.cell(row=1+index, column=1).value = str(key)
            sheet.cell(row=1+index, column=2).value = str(value)
       
        wb.save(REFERENCE_FILE)
        wb.close

    def save_in_db(self, data):
        certificate_data = copy.deepcopy(data)
        mongo.db.certificates.insert(certificate_data)
        #mongo.db.certificates.insert(data)

    def post1(self):
        data = {}
        data = {"hello": "there"}
        self.save_in_db(data)
        return jsonify({"hello": "there"})

    def get(self):
        return jsonify({"hello":"wassup!!"})

    def post(self):
        try:
            save_path = PDF_UPLOAD_DIRECTORY
            file = request.files['file']

            file_name = file.filename.replace(' ', '_')
            file_name_without_ext = os.path.basename(file_name).split('.')[0]
            file_name_without_ext = file_name_without_ext + "_" + str(uuid.uuid1())
            file_name = file_name_without_ext + path.splitext(file_name)[1]
            doc_dir_location = os.path.join(save_path, file_name_without_ext)
            if not os.path.exists(doc_dir_location):
                os.makedirs(doc_dir_location)
            file_location = os.path.join(doc_dir_location, file_name)
            file.save( file_location ) 
            
            result = read_scanned_pdf( file_location, doc_dir_location )
            print(f"--->{file_location}--->")
            #result = read_scanned_image( file_location, doc_dir_location)


            result['pdf_file_path'] = 'pdf_file/'   + file_name_without_ext
            result['excel_file_path'] = 'text_file/' + file_name_without_ext

            text_file_path = os.path.join(PDF_UPLOAD_DIRECTORY, file_name_without_ext, 'texts', 'stitched.txt')
            #text_file_path = os.path.join(PDF_UPLOAD_DIRECTORY, file_name_without_ext, 'texts', 'page-1.txt')
            print("text_file_path--->", text_file_path)
            with open( text_file_path ) as fp:
                contents = fp.readlines() 

            #self.parse_data(contents, result)
            parse_all_fields(contents, result) 
            self.save_in_db(result)
            self.update_excel_sheet(result, file.filename.replace(' ', '_'))
            return jsonify( {"data": result} )
            #return formulate_response(result, 200, "Successfully Extracted")

        except CustomClassifierException as e:
            print("1***ERROR***", e)
            logging.error("Error {} has occurred in controller".format(e))
            return e.response, e.http_code

        except Exception as e:
            print("2***ERROR***", e)
            logging.error("Error in service = {}".format(e), exc_info=True)
            return InternalServerErrorException(error_code=500,
                                                error_message="Data Extraction failed!").response, 500

        finally:
            logging.info("API Call Finished Successfully - 200")

    def create_template(self, template_path):
        sample_copy_path = "/Users/shravanc/flask/aditya_birla/ocr-pdf-aditya-malaysia/sample_copy/sample.xlsx"
        

        a = ['cp', sample_copy_path, template_path]
        template_file = os.path.join(template_path, 'sample.xlsx')
        res = subprocess.check_output(a)
        print(res)
        return template_file

